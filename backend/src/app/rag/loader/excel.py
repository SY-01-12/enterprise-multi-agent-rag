from __future__ import annotations

import openpyxl
from langchain_core.documents import Document


def load_xlsx(file_path: str) -> list[Document]:
    wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)

    parts: list[str] = []
    for sheet_name in wb.sheetnames:
        rows = [
            " | ".join(str(c) if c is not None else "" for c in row)
            for row in wb[sheet_name].iter_rows(values_only=True)
            if any(c is not None for c in row)
        ]
        if rows:
            parts.append(f"[工作表: {sheet_name}]\n" + "\n".join(rows))

    wb.close()

    if parts:
        return [Document(page_content="\n\n".join(parts), metadata={"source": file_path})]
    return []
